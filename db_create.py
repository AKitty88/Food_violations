import openpyxl
import sqlite3
from sqlite3 import Error

wb_ins = openpyxl.load_workbook("inspections.xlsx")
sheet_ins = wb_ins['inspections']

wb_viol = openpyxl.load_workbook("violations.xlsx")
sheet_viol = wb_viol['violations']

try:
    connection = sqlite3.connect("Food_violations.db")                  # creating a new database file (if it doesn't exist already)
    print("Food_violations.db created")    
except Error as e:
    connection.close()
    print(e)

cursor = connection.cursor()
cursor.execute("DROP TABLE IF EXISTS inspection;")
cursor.execute("DROP TABLE IF EXISTS violation;")
cursor.execute("DROP TABLE IF EXISTS previous_violation;")

sql_ins = """CREATE TABLE inspection (
    activity_date Date,
    employee_id char(10),    
    facility_address char(50),
    facility_city char(30),
    facility_id char(14),
    facility_name char(50),
    facility_state char(2),
    facility_zip char(14),
    grade char(1),
    owner_id char(15),
    owner_name char(50),
    pe_description char(50),
    program_element_pe number(4),
    program_name char(50),
    program_status char(10),
    record_id char(14), 
    score number(3),
    serial_number char(15) PRIMARY KEY,
    service_code number(3),
    service_description char(30));"""

sql_viol = """CREATE TABLE violation (
    points number(2),
    serial_number char(15),
    violation_code char(4),
    violation_description char(60),
    violation_status char(20),
    PRIMARY KEY (serial_number, violation_code),
    FOREIGN KEY (serial_number) REFERENCES inspection(serial_number),
    FOREIGN KEY (serial_number) REFERENCES previous_violation(serial_number));"""
          
sql_prev_viol = """CREATE TABLE previous_violation ( 
    facility_name char(50),
    facility_address char(50),
    facility_zip char(14),
    facility_city char(30),
    serial_number char(15) PRIMARY KEY,
    facility_id char(14));"""

cursor.execute(sql_ins)
cursor.execute(sql_prev_viol)
cursor.execute(sql_viol)

for row in sheet_ins.iter_rows(min_row=2):
    activity_date = row[0].value
    employee_id = row[1].value
    facility_address = row[2].value
    facility_city = row[3].value
    facility_id = row[4].value
    facility_name = row[5].value
    facility_state = row[6].value
    facility_zip = row[7].value
    grade = row[8].value
    owner_id = row[9].value
    owner_name = row[10].value
    pe_description = row[11].value
    program_element_pe = row[12].value
    program_name = row[13].value
    program_status = row[14].value
    record_id = row[15].value
    score = row[16].value
    serial_number = row[17].value
    service_code = row[18].value
    service_description = row[19].value
    
    cursor.execute("SELECT * FROM inspection WHERE serial_number=(?);", (serial_number,))
    found_ins = cursor.fetchall()
    
    if len(found_ins) == 0:
        cursor.execute("INSERT INTO inspection VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", (activity_date, employee_id, facility_address, facility_city, facility_id, facility_name, facility_state, facility_zip, grade, owner_id, owner_name, pe_description, program_element_pe, program_name,  program_status, record_id, score, serial_number, service_code, service_description))

for row in sheet_viol.iter_rows(min_row=2):
    points = row[0].value
    serial_number_v = row[1].value
    violation_code = row[2].value
    violation_description = row[3].value
    violation_status = row[4].value
    
    cursor.execute("SELECT * FROM violation WHERE serial_number=(?) AND violation_code=(?);", (serial_number_v, violation_code))
    found_viol = cursor.fetchall()
    
    if len(found_viol) == 0:
        cursor.execute("INSERT INTO violation VALUES(?, ?, ?, ?, ?);", (points, serial_number_v, violation_code, violation_description, violation_status))

# Task 2        
cursor.execute("SELECT DISTINCT facility_name, facility_address, facility_zip, facility_city, v.serial_number, facility_id FROM inspection i, violation v WHERE i.serial_number=v.serial_number ORDER BY facility_name;")
companies = cursor.fetchall()

for comp in companies:
    cursor.execute("SELECT * FROM previous_violation WHERE serial_number=(?);", (comp[4],))
    found_pr_viol = cursor.fetchall()
    
    if len(found_pr_viol) == 0:
        cursor.execute("INSERT INTO previous_violation VALUES(?, ?, ?, ?, ?, ?);", (comp[0], comp[1], comp[2], comp[3], comp[4], comp[5]))

cursor.execute("SELECT p.facility_id, p.facility_name, v.number \
               FROM previous_violation p, (SELECT DISTINCT COUNT(serial_number) AS number, serial_number FROM violation GROUP BY serial_number) v \
               WHERE v.serial_number = p.serial_number \
               ORDER BY v.number DESC;")
counts = cursor.fetchall()
print(counts)                                                                  # biggest: 27

cursor.execute("SELECT count(*) FROM inspection;")                              # 191.371
inspection = cursor.fetchall()
print(inspection)

cursor.execute("SELECT count(*) FROM violation;")                               # 905.885
violation = cursor.fetchall()
print(violation)

cursor.execute("SELECT count(*) FROM previous_violation;")                      # 186.899
previous_violation = cursor.fetchall()
print(previous_violation)

connection.commit()
cursor.close()
connection.close()
    





























