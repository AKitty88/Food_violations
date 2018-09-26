from openpyxl import Workbook
import sqlite3
from sqlite3 import Error

wb = Workbook()
filepath = "ViolationTypes.xlsx"
wb.save(filepath)

wb.create_sheet("Violations Types")
sheet = wb["Violations Types"]

try:
    connection = sqlite3.connect("Food_violations.db")
    print("Food_violations.db connected")
except Error as e:
    connection.close()
    print(e)

cursor = connection.cursor()
cursor.execute("SELECT violation_code FROM violation GROUP BY violation_code;")
violation_code = cursor.fetchall()
cursor.execute("SELECT violation_description FROM violation GROUP BY violation_code;")
violation_description = cursor.fetchall()
cursor.execute("SELECT COUNT(*) FROM violation GROUP BY violation_code;")
violation_counts = cursor.fetchall()

sheet['A1'] = "Code"
sheet['B1'] = "Description"
sheet['C1'] = "Count"

for i in range(2, len(violation_counts)):
    sheet.cell(row=i, column=1, value= str(violation_code[i]))
    sheet.cell(row=i, column=2, value= str(violation_description[i]))
    sheet.cell(row=i, column=3, value= str(violation_counts[i]))
    #sheet.cell(row=i, column=2, value="Dirty Floors")
    
wb.save('ViolationTypes.xlsx')
cursor.close()
connection.close()